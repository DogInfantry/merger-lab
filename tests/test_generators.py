"""Phase 5 verification: deal package assembly, memo HTML/PDF render,
Excel model structure + live-formula presence + engine tie-out cells.
Run: python tests/test_generators.py (also pytest-compatible).
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

from deal_package import build_deal_package
from excel_generator import generate_excel
from memo_generator import generate_memo, indian_group, inr_cr, render_memo_html
from precedent_db import get_connection, load_seed
from test_known_deal import ACQUIRER, TARGET, TERMS

OUT = Path(__file__).resolve().parent.parent / "samples" / "dev"


def _package():
    conn = get_connection(":memory:")
    load_seed(conn)
    return build_deal_package(
        ACQUIRER, TARGET, TERMS, codename="Project Anchor",
        strategic_rationale=["Consolidates a fragmented sector",
                             "Cross-sell into acquirer's distribution"],
        key_risks=["Synergy phasing slower than modeled",
                   "Integration cost overrun"],
        volatility=0.25, precedent_conn=conn, sector="Cement")


def test_indian_formatting():
    assert indian_group(628000) == "6,28,000"
    assert indian_group(1234) == "1,234"
    assert indian_group(-1234567) == "-12,34,567"
    assert inr_cr(628000) == "₹6,28,000 Cr"
    assert inr_cr(None) == "n/a"


def test_package_assembly():
    pkg = _package()
    # 100% stake still triggers SAST (>=25%) -> 3 scenario S&Us
    assert len(pkg.su_scenarios) == 3
    assert pkg.recommendation in ("PROCEED", "PROCEED WITH CONDITIONS", "DECLINE")
    assert pkg.recommendation_rationale
    assert pkg.collar is not None            # 50% stock + vol supplied
    assert pkg.sector_premium_percentile is not None
    # Toy deal at 100% acceptance: uses = 1200 + 312 + 1.5%(1512) = 1534.68
    assert abs(pkg.su.total_uses_cr - 1534.68) < 0.01


def test_memo_html_renders_all_sections():
    pkg = _package()
    html = render_memo_html(pkg)
    for needle in (
        "Investment Committee Memorandum", pkg.recommendation,
        "Transaction Summary", "Strategic Rationale", "Regulatory Stack",
        "acceptance", "RBI acquisition-finance compliance",
        "SEBI Takeover Code", "CCI merger control",
        "Accretion / Dilution Analysis", "Break-even synergies",
        "Contribution analysis", "Sensitivity",
        "Value Creation Assessment", "Risk Analysis",
        "Probability of accretion by Year 2", "<svg",
        "Exchange-ratio collar", "Appendix",
        "CONFIDENTIAL — ILLUSTRATIVE", "percentile of",
    ):
        assert needle in html, f"memo missing: {needle}"
    assert "{{" not in html and "{%" not in html, "unrendered Jinja left in memo"


def test_pdf_generation():
    pkg = _package()
    out = generate_memo(pkg, OUT / "project_anchor_memo.pdf")
    assert out.exists()
    if out.suffix == ".pdf":                  # engine available
        data = out.read_bytes()
        assert data[:5] == b"%PDF-", "not a valid PDF"
        assert len(data) > 30_000, "PDF suspiciously small"


def test_excel_model():
    pkg = _package()
    path = generate_excel(pkg, OUT / "project_anchor_model.xlsx")
    wb = load_workbook(path)                  # formulas preserved (not values)
    for tab in ("Assumptions", "Sources & Uses", "Regulatory", "PPA",
                "Pro-Forma P&L", "Accretion-Dilution", "Contribution",
                "Value Bridge", "Sensitivity", "Precedent Comps"):
        assert tab in wb.sheetnames, f"missing tab {tab}"

    su = wb["Sources & Uses"]
    assert su.cell(row=2, column=2).value.startswith("="), "S&U must be formulas"
    assert su.cell(row=11, column=1).value == "Balance check (must be 0)"
    ad = wb["Accretion-Dilution"]
    assert ad.cell(row=5, column=2).value.startswith("="), "accretion must be a formula"
    assert ad.cell(row=5, column=6).value.startswith("="), "engine tie-out column missing"
    pl = wb["Pro-Forma P&L"]
    assert all(pl.cell(row=9, column=j).value.startswith("=SUM") for j in (2, 3, 4))
    reg = wb["Regulatory"]
    assert "IF(" in reg.cell(row=2, column=4).value, "PASS/FAIL must be live IF()"

    # The one hardcode-only tab is Assumptions: value cells, blue font
    a = wb["Assumptions"]
    offer_cells = [c for row in a.iter_rows() for c in row
                   if c.value == "Offer price (₹)"]
    assert offer_cells and a.cell(row=offer_cells[0].row, column=2).value == 60.0


if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in fns:
        fn()
        print(f"PASS {fn.__name__}")
    print(f"\nall {len(fns)} generator tests OK")
