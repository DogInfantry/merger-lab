"""Excel model generator (openpyxl): a LINKED model, not a value dump.

methodology:
    Every input lives once on the Assumptions tab (blue font = hardcode,
    banker convention). Downstream tabs (Sources & Uses, PPA, Pro-Forma P&L,
    Accretion-Dilution, Value Bridge, Contribution, Regulatory) reference
    Assumptions and each other with REAL Excel formulas — change an
    assumption and the model recalculates. The Accretion-Dilution tab has a
    "Δ vs engine" column: formula result minus the frozen Python engine
    value; every delta must display 0 when opened, an in-file tie-out proof.
    Sensitivity grids and precedent comps are engine-computed values with
    conditional formatting (formulas there would need 99 full model copies).
    Formats: ₹ Cr with Indian digit grouping ([>=100000] custom section),
    % to 1dp, navy header fill, frozen panes.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from deal_package import DealPackage

NAVY = "070B14"
GOLD = "C9A84C"
BLUE = Font(color="0000CC")                      # hardcoded inputs
HDR_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN_TOP = Border(top=Side(style="thin"))
CR_FMT = '[>=100000]"₹ "##\\,##\\,##0" Cr";[<=-100000]"₹ -"##\\,##\\,##0" Cr";"₹ "#,##0.0" Cr"'
PCT_FMT = "+0.0%;-0.0%"
NUM_FMT = "#,##0.00"


def _sheet(wb: Workbook, title: str, headers: list[str], widths: list[int]):
    ws = wb.create_sheet(title)
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill, c.font = HDR_FILL, HDR_FONT
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


class _Assumptions:
    """Writes the Assumptions tab and remembers each input's cell address."""

    def __init__(self, wb: Workbook):
        self.ws = wb.create_sheet("Assumptions")
        self.ws.column_dimensions["A"].width = 44
        self.ws.column_dimensions["B"].width = 18
        self.row = 1
        self.cells: dict[str, str] = {}

    def section(self, name: str):
        c = self.ws.cell(row=self.row, column=1, value=name)
        c.fill, c.font = HDR_FILL, HDR_FONT
        self.ws.cell(row=self.row, column=2).fill = HDR_FILL
        self.row += 1

    def put(self, key: str, label: str, value, fmt: str | None = None):
        self.ws.cell(row=self.row, column=1, value=label)
        c = self.ws.cell(row=self.row, column=2, value=value)
        c.font = BLUE
        if fmt:
            c.number_format = fmt
        self.cells[key] = f"Assumptions!$B${self.row}"
        self.row += 1

    def __getitem__(self, key: str) -> str:
        return self.cells[key]


def generate_excel(pkg: DealPackage, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    a, t, terms = pkg.acquirer, pkg.target, pkg.terms

    # ---- Assumptions ---------------------------------------------------------
    A = _Assumptions(wb)
    A.section(f"ACQUIRER — {a.name}")
    A.put("a_price", "Share price (₹)", a.price, NUM_FMT)
    A.put("a_shares", "Shares outstanding (Cr)", a.shares_out_cr, NUM_FMT)
    A.put("a_ni", "Net income TTM (₹ Cr)", a.net_income_cr, CR_FMT)
    A.put("a_rev", "Revenue TTM (₹ Cr)", a.revenue_cr, CR_FMT)
    A.put("a_ebitda", "EBITDA TTM (₹ Cr)", a.ebitda_cr, CR_FMT)
    A.put("a_debt", "Total debt (₹ Cr)", a.total_debt_cr, CR_FMT)
    A.put("a_cash", "Cash (₹ Cr)", a.cash_cr, CR_FMT)
    A.put("a_book", "Book value (₹ Cr)", a.book_value_cr, CR_FMT)
    A.section(f"TARGET — {t.name}")
    A.put("t_price", "Undisturbed price (₹)", t.price, NUM_FMT)
    A.put("t_shares", "Diluted shares (Cr)", t.shares_out_cr, NUM_FMT)
    A.put("t_ni", "Net income TTM (₹ Cr)", t.net_income_cr, CR_FMT)
    A.put("t_rev", "Revenue TTM (₹ Cr)", t.revenue_cr, CR_FMT)
    A.put("t_ebitda", "EBITDA TTM (₹ Cr)", t.ebitda_cr, CR_FMT)
    A.put("t_debt", "Total debt (₹ Cr)", t.total_debt_cr, CR_FMT)
    A.put("t_book", "Book value (₹ Cr)", t.book_value_cr, CR_FMT)
    A.section("DEAL TERMS")
    A.put("offer", "Offer price (₹)", terms.offer_price, NUM_FMT)
    A.put("stake", "Negotiated stake %", terms.stake_pct / 100, "0.0%")
    A.put("pct_cash", "% cash consideration", terms.pct_cash / 100, "0.0%")
    A.put("pct_stock", "% stock consideration", terms.pct_stock / 100, "0.0%")
    A.put("pct_debt", "% of cash needs from new debt",
          terms.pct_new_debt_of_cash_portion / 100, "0.0%")
    A.put("rate", "New debt interest rate", terms.debt_interest_rate, "0.00%")
    A.put("yield", "Cash yield foregone", terms.cash_yield_foregone, "0.00%")
    A.put("syn", "Run-rate synergies (₹ Cr)", terms.synergies_annual, CR_FMT)
    A.put("ph1", "Synergy phase-in Y1", terms.synergy_phase_in[0], "0%")
    A.put("ph2", "Synergy phase-in Y2", terms.synergy_phase_in[1], "0%")
    A.put("ph3", "Synergy phase-in Y3", terms.synergy_phase_in[2], "0%")
    A.put("integ", "Integration costs Y1 (₹ Cr)", terms.integration_costs, CR_FMT)
    A.put("writeup", "Intangible write-up % of excess",
          terms.intangible_writeup_pct, "0.0%")
    A.put("life", "Intangible life (years)", terms.intangible_life_years)
    A.put("tax", "Tax rate", terms.tax_rate, "0.00%")
    A.put("fees", "Transaction fees %", terms.fees_pct, "0.00%")
    A.put("wacc", "WACC", pkg.wacc, "0.00%")
    A.put("oo_accept", "Open-offer acceptance assumption",
          pkg.acceptance_assumption_pct / 100, "0%")
    A.section("CONVENTION")
    A.ws.cell(row=A.row, column=1,
              value="Blue font = hardcoded input. All other tabs are formulas.")

    # ---- Sources & Uses (scenario columns, live formulas) --------------------
    accs = list(pkg.su_scenarios.keys())
    su_ws = _sheet(wb, "Sources & Uses",
                   ["₹ Cr"] + [f"{acc:.0f}% acceptance" for acc in accs],
                   [34] + [18] * len(accs))
    su_cell: dict[tuple[str, int], str] = {}

    def su_row(r: int, label: str, formula_for):
        su_ws.cell(row=r, column=1, value=label)
        for j, acc in enumerate(accs, start=2):
            c = su_ws.cell(row=r, column=j, value=formula_for(acc, get_column_letter(j)))
            c.number_format = CR_FMT
            su_cell[(label, acc)] = f"'Sources & Uses'!{get_column_letter(j)}{r}"

    eq = f"{A['offer']}*{A['t_shares']}*{A['stake']}"
    su_row(2, "Equity purchase (negotiated)", lambda acc, L: f"={eq}")
    su_row(3, "Open offer consideration",
           lambda acc, L: f"=26%*{A['t_shares']}*{A['offer']}*{acc / 100}")
    su_row(4, "Refinance target debt",
           lambda acc, L: f"={A['t_debt']}*{1 if terms.refinance_target_debt else 0}")
    su_row(5, "Transaction fees", lambda acc, L: f"={A['fees']}*({L}2+{L}3)")
    su_row(6, "TOTAL USES", lambda acc, L: f"=SUM({L}2:{L}5)")
    su_row(7, "New acquirer stock", lambda acc, L: f"={A['pct_stock']}*{L}2")
    su_row(8, "New bank debt",
           lambda acc, L: f"={A['pct_debt']}*({A['pct_cash']}*{L}2+{L}3+{L}4+{L}5)")
    su_row(9, "Balance-sheet cash",
           lambda acc, L: f"=({A['pct_cash']}*{L}2+{L}3+{L}4+{L}5)-{L}8")
    su_row(10, "TOTAL SOURCES", lambda acc, L: f"=SUM({L}7:{L}9)")
    su_row(11, "Balance check (must be 0)", lambda acc, L: f"={L}10-{L}6")
    su_row(12, "New shares issued (Cr)", lambda acc, L: f"={L}7/{A['a_price']}")
    su_ws.cell(row=12, column=2).number_format = NUM_FMT
    for row in (6, 10):
        for j in range(1, len(accs) + 2):
            su_ws.cell(row=row, column=j).font = BOLD
            su_ws.cell(row=row, column=j).border = THIN_TOP
    for j in range(2, len(accs) + 2):
        su_ws.cell(row=12, column=j).number_format = NUM_FMT

    acc0 = pkg.acceptance_assumption_pct
    SU = lambda label: su_cell[(label, acc0)]  # scenario column used downstream

    # ---- PPA -----------------------------------------------------------------
    ppa_ws = _sheet(wb, "PPA", ["Item", "₹ Cr"], [40, 18])
    owned = f"({A['stake']}+26%*{A['oo_accept']})"
    ppa_rows = [
        ("Equity invested (negotiated + open offer)",
         f"={SU('Equity purchase (negotiated)')}+{SU('Open offer consideration')}"),
        ("Owned share of target book value", f"={owned}*{A['t_book']}"),
        ("Excess over book", "=B2-B3"),
        ("Intangible write-up", f"=MAX(0,{A['writeup']}*B4)"),
        ("Deferred tax liability", f"={A['tax']}*B5"),
        ("Goodwill", "=B4-B5+B6"),
        ("Incremental D&A per year", f"=B5/{A['life']}"),
    ]
    for i, (lbl, f) in enumerate(ppa_rows, start=2):
        ppa_ws.cell(row=i, column=1, value=lbl)
        c = ppa_ws.cell(row=i, column=2, value=f)
        c.number_format = CR_FMT

    # ---- Pro-Forma P&L -------------------------------------------------------
    pl_ws = _sheet(wb, "Pro-Forma P&L", ["₹ Cr", "Year 1", "Year 2", "Year 3"],
                   [38, 16, 16, 16])
    at = f"(1-{A['tax']})"
    phases = [A["ph1"], A["ph2"], A["ph3"]]
    pl_rows = [
        ("Acquirer net income", [f"={A['a_ni']}"] * 3),
        ("Owned target net income", [f"={owned}*{A['t_ni']}"] * 3),
        ("Synergies (phased, after-tax)",
         [f"={A['syn']}*{ph}*{at}" for ph in phases]),
        ("New interest expense (after-tax)",
         [f"=-{SU('New bank debt')}*{A['rate']}*{at}"] * 3),
        ("Foregone cash yield (after-tax)",
         [f"=-{SU('Balance-sheet cash')}*{A['yield']}*{at}"] * 3),
        ("Incremental D&A (after-tax)", ["=-PPA!B8*" + at] * 3),
        ("Integration costs (after-tax, Y1)",
         [f"=-{A['integ']}*{at}" if terms.include_integration_costs else "=0",
          "=0", "=0"]),
    ]
    for i, (lbl, fs) in enumerate(pl_rows, start=2):
        pl_ws.cell(row=i, column=1, value=lbl)
        for j, f in enumerate(fs, start=2):
            pl_ws.cell(row=i, column=j, value=f).number_format = CR_FMT
    pl_ws.cell(row=9, column=1, value="COMBINED NET INCOME").font = BOLD
    for j, L in enumerate("BCD", start=2):
        c = pl_ws.cell(row=9, column=j, value=f"=SUM({L}2:{L}8)")
        c.number_format, c.font, c.border = CR_FMT, BOLD, THIN_TOP

    # ---- Accretion-Dilution (with engine tie-out) ----------------------------
    ad_ws = _sheet(wb, "Accretion-Dilution",
                   ["Item", "Year 1", "Year 2", "Year 3", "", "Δ vs engine (Y1)"],
                   [34, 15, 15, 15, 3, 18])
    ad_ws.cell(row=2, column=1, value="Pro-forma shares (Cr)")
    for j, L in enumerate("BCD", start=2):
        ad_ws.cell(row=2, column=j,
                   value=f"={A['a_shares']}+{SU('New shares issued (Cr)')}"
                   ).number_format = NUM_FMT
    ad_ws.cell(row=3, column=1, value="Pro-forma EPS (₹)")
    for j, L in enumerate("BCD", start=2):
        ad_ws.cell(row=3, column=j,
                   value=f"='Pro-Forma P&L'!{L}9/{get_column_letter(j)}2"
                   ).number_format = NUM_FMT
    ad_ws.cell(row=4, column=1, value="Standalone acquirer EPS (₹)")
    for j in range(2, 5):
        ad_ws.cell(row=4, column=j,
                   value=f"={A['a_ni']}/{A['a_shares']}").number_format = NUM_FMT
    ad_ws.cell(row=5, column=1, value="Accretion / (dilution)").font = BOLD
    for j, L in enumerate("BCD", start=2):
        c = ad_ws.cell(row=5, column=j, value=f"={L}3/{L}4-1")
        c.number_format, c.font = PCT_FMT, BOLD
    # engine tie-out: formula minus frozen Python value must show 0
    eng = {2: pkg.ad.years[0].pf_shares_cr, 3: pkg.ad.years[0].pf_eps,
           4: pkg.ad.years[0].standalone_eps,
           5: pkg.ad.years[0].accretion_pct / 100}
    for r, v in eng.items():
        ad_ws.cell(row=r, column=6, value=f"=B{r}-{v!r}").number_format = "0.000000"
    ad_ws.cell(row=7, column=1, value="Break-even synergies Y1 (₹ Cr)")
    ad_ws.cell(row=7, column=2,
               value=f"=(B4*B2-('Pro-Forma P&L'!B9-{A['syn']}*{A['ph1']}*{at}))"
                     f"/({A['ph1']}*{at})").number_format = CR_FMT

    # ---- Regulatory ----------------------------------------------------------
    reg_ws = _sheet(wb, "Regulatory", ["Check", "Value", "Threshold", "Status"],
                    [46, 16, 16, 12])
    acq_val = f"({SU('Equity purchase (negotiated)')}+{SU('Open offer consideration')})"
    reg_rows = [
        ("RBI: bank debt ≤ 75% of acquisition value",
         f"={SU('New bank debt')}/{acq_val}", "0.0%", "≤ 75%", "=IF(B2<=75%,\"PASS\",\"FAIL\")"),
        ("RBI: equity contribution ≥ 25%",
         f"=1-{SU('New bank debt')}/{SU('TOTAL SOURCES')}", "0.0%", "≥ 25%",
         "=IF(B3>=25%,\"PASS\",\"FAIL\")"),
        ("RBI: pro-forma consolidated D/E ≤ 3.0x",
         f"=({A['a_debt']}+{SU('New bank debt')}"
         f"+{A['t_debt']}*{0 if terms.refinance_target_debt else 1})"
         f"/({A['a_book']}+{SU('New acquirer stock')})", "0.00\"x\"", "≤ 3.0x",
         "=IF(B4<=3,\"PASS\",\"FAIL\")"),
        ("RBI: acquirer net worth ≥ ₹500 Cr", f"={A['a_book']}", CR_FMT, "≥ ₹500 Cr",
         "=IF(B5>=500,\"PASS\",\"FAIL\")"),
    ]
    for i, (lbl, f, fmt, thr, status) in enumerate(reg_rows, start=2):
        reg_ws.cell(row=i, column=1, value=lbl)
        reg_ws.cell(row=i, column=2, value=f).number_format = fmt
        reg_ws.cell(row=i, column=3, value=thr)
        reg_ws.cell(row=i, column=4, value=status)
    r = len(reg_rows) + 3
    reg_ws.cell(row=r, column=1, value="SEBI SAST").font = BOLD
    reg_ws.cell(row=r + 1, column=1, value=pkg.sast.narrative)
    for i, s in enumerate(pkg.sast.scenarios):
        reg_ws.cell(row=r + 2 + i, column=1, value=s.narrative)
    r += 3 + len(pkg.sast.scenarios)
    reg_ws.cell(row=r, column=1, value="CCI").font = BOLD
    reg_ws.cell(row=r + 1, column=1,
                value=pkg.sast.cci_note if pkg.sast.cci_approval_required
                else "Below ₹2,000 Cr deal-value threshold.")

    # ---- Contribution --------------------------------------------------------
    con_ws = _sheet(wb, "Contribution", ["Metric", "Acquirer %", "Target %"],
                    [26, 14, 14])
    con_rows = [
        ("Revenue", A["a_rev"], A["t_rev"]),
        ("EBITDA", A["a_ebitda"], A["t_ebitda"]),
        ("Net income", A["a_ni"], A["t_ni"]),
    ]
    for i, (lbl, af, tf) in enumerate(con_rows, start=2):
        con_ws.cell(row=i, column=1, value=lbl)
        con_ws.cell(row=i, column=2, value=f"={af}/({af}+{tf})").number_format = "0.0%"
        con_ws.cell(row=i, column=3, value=f"={tf}/({af}+{tf})").number_format = "0.0%"
    con_ws.cell(row=5, column=1, value="Pro-forma ownership")
    own_t = f"{SU('New shares issued (Cr)')}/({A['a_shares']}+{SU('New shares issued (Cr)')})"
    con_ws.cell(row=5, column=2, value=f"=1-{own_t}").number_format = "0.0%"
    con_ws.cell(row=5, column=3, value=f"={own_t}").number_format = "0.0%"
    if pkg.contribution_flag:
        con_ws.cell(row=7, column=1, value=pkg.contribution_flag)

    # ---- Value Bridge --------------------------------------------------------
    vb_ws = _sheet(wb, "Value Bridge", ["Item", "₹ Cr"], [44, 18])
    vb_rows = [
        ("PV of after-tax synergies (perpetuity at WACC)",
         f"={A['syn']}*(1-{A['tax']})/{A['wacc']}"),
        ("Control premium paid",
         f"=({A['offer']}-{A['t_price']})*{A['t_shares']}*{owned}"),
        ("Net value created / (destroyed)", "=B2-B3"),
        ("Incremental NOPAT (owned NI + after-tax synergies)",
         f"={owned}*{A['t_ni']}+{A['syn']}*(1-{A['tax']})"),
        ("Invested capital (total uses)", f"={SU('TOTAL USES')}"),
    ]
    for i, (lbl, f) in enumerate(vb_rows, start=2):
        vb_ws.cell(row=i, column=1, value=lbl)
        vb_ws.cell(row=i, column=2, value=f).number_format = CR_FMT
    vb_ws.cell(row=7, column=1, value="Incremental ROIC")
    vb_ws.cell(row=7, column=2, value="=B5/B6").number_format = "0.0%"
    vb_ws.cell(row=8, column=1, value="WACC")
    vb_ws.cell(row=8, column=2, value=f"={A['wacc']}").number_format = "0.0%"
    if pkg.value_bridge.mechanical_accretion_warning:
        c = vb_ws.cell(row=10, column=1,
                       value="WARNING: accretive but value-destructive (P/E arbitrage)")
        c.font = Font(bold=True, color="B02A1E")

    # ---- Sensitivity (values + heat conditional formatting) ------------------
    sen_ws = _sheet(wb, "Sensitivity", [""], [24])
    def dump_grid(ws, df: pd.DataFrame, start_row: int, title: str) -> int:
        ws.cell(row=start_row, column=1, value=title).font = BOLD
        r0 = start_row + 1
        ws.cell(row=r0, column=1, value=df.index.name)
        for j, col in enumerate(df.columns, start=2):
            c = ws.cell(row=r0, column=j, value=col)
            c.fill, c.font = HDR_FILL, HDR_FONT
            ws.column_dimensions[get_column_letter(j)].width = 10
        for i, (idx, row) in enumerate(df.iterrows(), start=r0 + 1):
            ws.cell(row=i, column=1, value=idx).font = BOLD
            for j, v in enumerate(row, start=2):
                ws.cell(row=i, column=j, value=float(v)).number_format = "+0.00;-0.00"
        rng = (f"B{r0 + 1}:{get_column_letter(len(df.columns) + 1)}"
               f"{r0 + len(df)}")
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="C0392B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="1D7A4A"))
        return r0 + len(df) + 2
    nxt = dump_grid(sen_ws, pkg.grid_premium_synergies, 2,
                    "Year-1 accretion %: premium (rows) × synergies ₹ Cr (cols) — engine values")
    dump_grid(sen_ws, pkg.grid_cash_premium, nxt,
              "Year-1 accretion %: % cash (rows) × premium (cols) — engine values")

    # ---- Precedent Comps -----------------------------------------------------
    if pkg.sector_comps is not None and not pkg.sector_comps.empty:
        cols = ["announce_date", "acquirer", "target", "deal_value_cr",
                "consideration_type", "offer_premium_pct", "ev_ebitda_multiple",
                "status", "notes", "source_url"]
        pc_ws = _sheet(wb, "Precedent Comps", cols,
                       [12, 26, 26, 14, 12, 12, 12, 11, 30, 45])
        for i, row in enumerate(pkg.sector_comps[cols].itertuples(index=False),
                                start=2):
            for j, v in enumerate(row, start=1):
                pc_ws.cell(row=i, column=j, value=v)

    wb.save(path)
    return path
